# Streamlit app (orientador + alumnado) para "PONDERACIONES.xlsx"
# - Panel orientador: análisis global de qué asignaturas ponderan más (0,20 / 0,10) por ámbito
# - Herramienta alumnado: el alumno elige 4 asignaturas (solo entre las que ponderan 0,10 o 0,20) y ve ponderaciones
#
# Ejecuta: python -m streamlit run app_orientador_centro.py

import re
from collections import Counter

import pandas as pd
import openpyxl
import streamlit as st

EXCEL_PATH = "PONDERACIONES.xlsx"

st.set_page_config(
    page_title="Ponderaciones · Panel de Orientación",
    layout="wide",
    page_icon="🎓",
)

# ---------------------------
# Estilos (sobrio y legible)
# ---------------------------
st.markdown(
    """
<style>
.main {
  background: radial-gradient(circle at 10% 0%, #e0f2fe 0%, transparent 35%),
              radial-gradient(circle at 90% 10%, #ede9fe 0%, transparent 40%),
              linear-gradient(135deg, #f8fafc 0%, #eef2ff 100%);
}
.block-container { padding-top: 1.4rem; }
h1 { color: #0f172a; font-weight: 900; letter-spacing: -0.02em; }
h2, h3 { color: #0b3b8f; }
.stTextInput>div>div>input,
.stMultiSelect>div>div,
.stSelectbox>div>div { border-radius: 12px !important; }
.hr {
  height: 1px;
  background: linear-gradient(90deg, transparent, rgba(148, 163, 184, 0.7), transparent);
  margin: 14px 0;
}
.small { color: rgba(15,23,42,0.70); font-weight: 600; font-size: 0.95rem; }
.badge {
  display: inline-block;
  padding: 2px 10px;
  border-radius: 999px;
  background: rgba(59,130,246,0.12);
  border: 1px solid rgba(59,130,246,0.18);
  font-weight: 800;
  color: #1e3a8a;
}
</style>
""",
    unsafe_allow_html=True,
)

# -----------------
# Carga del Excel
# -----------------
@st.cache_data(show_spinner=False)
def load_data(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Asignaturas (fila 2, columnas C..)
    subjects = []
    for c in range(3, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v is None:
            break
        subjects.append(str(v).strip())

    # Detectar universidades frecuentes entre paréntesis (heurística)
    deg_cells = [ws.cell(r, 2).value for r in range(3, ws.max_row + 1) if ws.cell(r, 2).value]
    paren_counter = Counter()
    for s in deg_cells:
        for g in re.findall(r"\(([^()]*)\)", str(s)):
            paren_counter[g.strip()] += 1

    uni_candidates = set()
    for g, cnt in paren_counter.items():
        if (
            cnt >= 5
            and len(g) <= 14
            and re.fullmatch(r"[A-Z0-9/ .-]+", g)
            and (not g.isdigit())
            and (not re.search(r"OPCI|PRESEN|SEMIP|CONJ|INTER|CENTRO|\*", g, re.I))
        ):
            uni_candidates.add(g)
    # fallback de universidades típicas (por si el excel tiene pocas repeticiones)
    uni_candidates |= {"UAM", "UCM", "URJC", "UPM", "UC3M", "UAH"}

    rows = []
    for r in range(3, ws.max_row + 1):
        full = ws.cell(r, 2).value
        if full is None:
            continue

        s = str(full).strip()
        groups = [g.strip() for g in re.findall(r"\(([^()]*)\)", s)]

        uni = None
        for g in groups:
            if g in uni_candidates:
                uni = g
                break
        if uni is None:
            for g in groups:
                for cand in uni_candidates:
                    if cand in g:
                        uni = cand
                        break
                if uni:
                    break

        grado = re.sub(r"\s*\([^()]*\)\s*", " ", s).strip()
        grado = re.sub(r"\s+", " ", grado)

        coeffs = {}
        for i, subj in enumerate(subjects, start=3):
            v = ws.cell(r, i).value
            if v is None or v == "":
                coeffs[subj] = None
            else:
                try:
                    coeffs[subj] = float(str(v).replace(",", "."))
                except Exception:
                    coeffs[subj] = None

        row = {"grado": grado, "universidad": uni, "titulo_excel": s}
        row.update(coeffs)
        rows.append(row)

    df = pd.DataFrame(rows)
    return df, subjects


def style_weights(df: pd.DataFrame):
    # Colorea celdas por ponderación (pensado para 0,10 y 0,20)
    def cell_style(v):
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return "background-color: rgba(148,163,184,0.08); color: rgba(15,23,42,0.55);"
            v = float(v)
        except Exception:
            return ""
        if v == 0:
            return "background-color: rgba(148,163,184,0.08); color: rgba(15,23,42,0.55);"
        if abs(v - 0.2) < 1e-6:
            return "background-color: rgba(16,185,129,0.22); color: #065f46; font-weight: 800;"
        if abs(v - 0.1) < 1e-6:
            return "background-color: rgba(245,158,11,0.22); color: #7c2d12; font-weight: 800;"
        if v > 0:
            return "background-color: rgba(59,130,246,0.16); color: #1e3a8a; font-weight: 800;"
        return ""

    cols = [c for c in df.columns if c not in ("Grado", "Universidad")]
    return df.style.applymap(cell_style, subset=cols).format({c: "{:.2f}" for c in cols})




# -----------------
# Tabla con columnas fijas (Grado/Universidad)
# -----------------
def render_sticky_table(df: pd.DataFrame, sticky_cols: int = 2, height_px: int = 520):
    """
    Renderiza una tabla HTML con las primeras `sticky_cols` columnas fijas (sticky).
    Útil cuando hay muchas asignaturas y se quiere mantener "Grado" y "Universidad" visibles.
    """
    if df.empty:
        st.info("No hay datos para mostrar.")
        return

    html = df.to_html(index=False, escape=True)

    st.markdown(
        f"""
<style>
.table-wrap {{
  max-height: {height_px}px;
  overflow: auto;
  border-radius: 14px;
  border: 1px solid rgba(148,163,184,0.35);
  background: rgba(255,255,255,0.85);
}}
.table-wrap table {{
  border-collapse: separate;
  border-spacing: 0;
  width: 100%;
  font-size: 0.95rem;
}}
.table-wrap th, .table-wrap td {{
  padding: 8px 10px;
  border-bottom: 1px solid rgba(148,163,184,0.25);
  white-space: nowrap;
}}
.table-wrap th {{
  position: sticky;
  top: 0;
  z-index: 5;
  background: rgba(241,245,249,0.98);
  color: #0f172a;
  font-weight: 900;
}}
/* columnas fijas */
.table-wrap th:nth-child(1), .table-wrap td:nth-child(1) {{
  position: sticky;
  left: 0;
  z-index: 6;
  background: rgba(248,250,252,0.98);
}}
.table-wrap th:nth-child(2), .table-wrap td:nth-child(2) {{
  position: sticky;
  left: 190px; /* ancho aproximado de la 1ª columna */
  z-index: 6;
  background: rgba(248,250,252,0.98);
}}
/* sombra separadora después de columnas fijas */
.table-wrap td:nth-child(2), .table-wrap th:nth-child(2) {{
  box-shadow: 8px 0 10px -8px rgba(2,6,23,0.35);
}}
/* colores de ponderación */
.w20 {{ background: rgba(16,185,129,0.22); color: #065f46; font-weight: 900; padding: 2px 6px; border-radius: 8px; }}
.w10 {{ background: rgba(245,158,11,0.22); color: #7c2d12; font-weight: 900; padding: 2px 6px; border-radius: 8px; }}
</style>
""",
        unsafe_allow_html=True,
    )

    html2 = html
    html2 = re.sub(r">(0\.20|0,20)<", r'><span class="w20">\\1</span><', html2)
    html2 = re.sub(r">(0\.10|0,10)<", r'><span class="w10">\\1</span><', html2)

    st.markdown(f'<div class="table-wrap">{html2}</div>', unsafe_allow_html=True)
def subjects_that_ponder_01_02(df_scope: pd.DataFrame, subjects: list[str]) -> list[str]:
    cand = []
    for s in subjects:
        col = pd.to_numeric(df_scope[s], errors="coerce").round(2)
        if ((col == 0.10) | (col == 0.20)).any():
            cand.append(s)
    return cand


def rank_subjects(df_scope: pd.DataFrame, subjects: list[str]) -> pd.DataFrame:
    # Devuelve un dataframe con métricas útiles para orientadores
    rows = []
    n = len(df_scope)
    for s in subjects:
        col = pd.to_numeric(df_scope[s], errors="coerce").round(2)
        cnt_02 = int((col == 0.20).sum())
        cnt_01 = int((col == 0.10).sum())
        cnt_any = int(((col == 0.20) | (col == 0.10)).sum())
        pct_02 = (cnt_02 / n) * 100 if n else 0.0
        pct_any = (cnt_any / n) * 100 if n else 0.0
        rows.append(
            {
                "Asignatura": s,
                "Veces 0,20": cnt_02,
                "Veces 0,10": cnt_01,
                "Veces (0,10 o 0,20)": cnt_any,
                "% con 0,20": pct_02,
                "% con (0,10 o 0,20)": pct_any,
            }
        )
    out = pd.DataFrame(rows)
    out = out.sort_values(["Veces 0,20", "Veces (0,10 o 0,20)", "Asignatura"], ascending=[False, False, True])
    return out.reset_index(drop=True)


# -----------
# UI
# -----------
st.title("🎓 Ponderaciones: panel para orientación y departamentos")
st.markdown(
    '<div class="small">Pensada para orientadores/jefaturas: filtra el ámbito (universidades y grados) '
    'y analiza qué asignaturas ponderan más en general. Incluye una vista “alumnado”.</div>',
    unsafe_allow_html=True,
)

df, SUBJECTS = load_data(EXCEL_PATH)

tabs = st.tabs(["Panel orientador", "Vista alumnado"])

# =========================
# TAB 1 — PANEL ORIENTADOR
# =========================
with tabs[0]:
    st.subheader("🎛️ Ámbito de análisis")
    with st.container(border=True):
        c1, c2, c3 = st.columns([2.3, 2.3, 1.4])

        with c1:
            # Selección explícita de grados (sin búsqueda por texto)
            # Se rellena tras aplicar el filtro de universidades
            st.markdown("**📚 Grados (selecciona varios)**")

        with c2:
            unis = sorted([u for u in df["universidad"].dropna().unique().tolist()])
            b1, b2 = st.columns(2)
            with b1:
                if st.button("🏛️ Todas", key="ori_unis_all"):
                    st.session_state["ori_unis"] = unis
            with b2:
                if st.button("🧹 Ninguna", key="ori_unis_none"):
                    st.session_state["ori_unis"] = []
            if "ori_unis" not in st.session_state:
                st.session_state["ori_unis"] = []
            selected_unis = st.multiselect("Universidades", options=unis, key="ori_unis")

        with c3:
            st.markdown('<span class="badge">Consejo</span> Elige universidades y luego grados.', unsafe_allow_html=True)

    df_f = df.copy()
    if selected_unis:
        df_f = df_f[df_f["universidad"].isin(selected_unis)]

    # Lista de grados disponible tras filtrar universidades
    grados_ori = sorted(df_f["grado"].dropna().unique().tolist())

    # Botones de selección rápida para grados
    g1, g2 = st.columns(2)
    with g1:
        if st.button("📚 Todos los grados", key="ori_grados_all"):
            st.session_state["ori_grados"] = grados_ori
    with g2:
        if st.button("🧹 Ninguno", key="ori_grados_none"):
            st.session_state["ori_grados"] = []

    if "ori_grados" not in st.session_state:
        st.session_state["ori_grados"] = []

    selected_grados = st.multiselect(
        "Grados",
        options=grados_ori,
        key="ori_grados",
    )

    if selected_grados:
        df_f = df_f[df_f["grado"].isin(selected_grados)]

    st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

    st.subheader("📌 Indicadores globales")
    if df_f.empty:
        st.info("No hay resultados con los filtros actuales.")
        st.stop()

    cand = subjects_that_ponder_01_02(df_f, SUBJECTS)
    metrics = rank_subjects(df_f, cand)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Filas (grado+uni) en el ámbito", f"{len(df_f)}")
    m2.metric("Asignaturas que ponderan (0,10/0,20)", f"{len(cand)}")
    m3.metric("Top asignatura por 0,20", metrics.iloc[0]["Asignatura"] if not metrics.empty else "—")
    m4.metric("% máx. con 0,20", f"{metrics.iloc[0]['% con 0,20']:.1f}%" if not metrics.empty else "—")

    st.markdown("### 🥇 Top asignaturas que más ponderan (0,20)")
    st.dataframe(
        metrics.head(20).style.format(
            {"% con 0,20": "{:.1f}%", "% con (0,10 o 0,20)": "{:.1f}%"}
        ),
        use_container_width=True,
        hide_index=True,
    )

    st.markdown("### 🔎 Explorar una asignatura")
    subj = st.selectbox("Selecciona asignatura", options=cand)
    if subj:
        tmp = df_f[["grado", "universidad", subj]].copy()
        tmp[subj] = pd.to_numeric(tmp[subj], errors="coerce").fillna(0).round(2)
        tmp = tmp[tmp[subj].isin([0.10, 0.20])].sort_values([subj, "grado"], ascending=[False, True])
        tmp = tmp.rename(columns={"grado": "Grado", "universidad": "Universidad", subj: "Ponderación"})
        st.dataframe(tmp, use_container_width=True, hide_index=True)

    st.markdown("### 🧾 Matriz rápida (opcional)")
    with st.expander("Ver matriz grados/universidades × asignaturas (puede ser grande)"):
        # Para no petar la pantalla, permitimos elegir 6-10 asignaturas a mostrar
        show_subjs = st.multiselect("Asignaturas a mostrar (recomendado 6–10)", options=cand, default=cand[:8])
        if show_subjs:
            mat = df_f[["grado", "universidad"] + show_subjs].copy()
            mat = mat.rename(columns={"grado": "Grado", "universidad": "Universidad"})
            # normalizar a 0 si NaN
            for s in show_subjs:
                mat[s] = pd.to_numeric(mat[s], errors="coerce").fillna(0).round(2)
            st.dataframe(style_weights(mat), use_container_width=True, hide_index=True)



    st.markdown("### 🗂️ Datos completos (Excel) del ámbito filtrado")
    with st.expander("Ver todas las columnas del Excel para los grados/universidades filtrados"):
        all_cols = ["grado", "universidad", "titulo_excel"] + SUBJECTS
        st.dataframe(df_f[all_cols], use_container_width=True, hide_index=True)

# =====================
# TAB 2 — VISTA ALUMNADO
# =====================
with tabs[1]:
    st.subheader("👩‍🎓 Herramienta para alumnado (selección de 4)")
    st.caption("Pensada para usar en tutoría: el alumno elige 4 asignaturas y ve su ponderación por grado/universidad.")

    with st.container(border=True):
        a1, a2 = st.columns([2.2, 2.2])
        with a1:
            q2 = st.text_input("🔎 Buscar grado (texto)", value="", key="stu_q")
        with a2:
            unis2 = sorted([u for u in df["universidad"].dropna().unique().tolist()])
            b1, b2 = st.columns(2)
            with b1:
                if st.button("🏛️ Todas", key="stu_unis_all"):
                    st.session_state["stu_unis"] = unis2
            with b2:
                if st.button("🧹 Ninguna", key="stu_unis_none"):
                    st.session_state["stu_unis"] = []
            if "stu_unis" not in st.session_state:
                st.session_state["stu_unis"] = []
            selected_unis2 = st.multiselect("Universidades", options=unis2, key="stu_unis")

    df2 = df.copy()
    if selected_unis2:
        df2 = df2[df2["universidad"].isin(selected_unis2)]
    if q2.strip():
        df2 = df2[df2["grado"].str.lower().str.contains(q2.strip().lower(), na=False)]

    grados2 = sorted(df2["grado"].dropna().unique().tolist())
    selected_grados2 = st.multiselect("📚 Selecciona grados", grados2, key="stu_grados")

    if not selected_grados2:
        st.info("Selecciona al menos un grado para que solo salgan asignaturas que ponderan.")
        st.stop()

    df_scope = df2[df2["grado"].isin(selected_grados2)]
    cand2 = subjects_that_ponder_01_02(df_scope, SUBJECTS)
    metrics2 = rank_subjects(df_scope, cand2)
    ordered2 = metrics2["Asignatura"].tolist()  # orden por “mejores”

    if "stu_chosen" in st.session_state:
        st.session_state["stu_chosen"] = [s for s in st.session_state["stu_chosen"] if s in ordered2]

    st.markdown('<div class="hr"></div>', unsafe_allow_html=True)
    st.markdown("### ✅ Paso 2: elige 4 asignaturas (solo 0,10 / 0,20)")

    with st.container(border=True):
        b1, b2, b3 = st.columns([1.6, 1.1, 2.3])
        with b1:
            if st.button("✨ Autoseleccionar 4 (mejores)", key="stu_auto4"):
                st.session_state["stu_chosen"] = ordered2[:4]
        with b2:
            if st.button("🧹 Limpiar", key="stu_clear4"):
                st.session_state["stu_chosen"] = []
        with b3:
            st.caption("Criterio: más veces 0,20 en el ámbito; luego más veces (0,10/0,20).")

    if "stu_chosen" not in st.session_state:
        st.session_state["stu_chosen"] = []

    chosen = st.multiselect("Asignaturas (elige 4)", options=ordered2, key="stu_chosen")

    if len(chosen) != 4:
        st.info("Elige exactamente 4 asignaturas para ver la tabla.")
        st.stop()

    st.markdown("### 📊 Ponderación de las 4 asignaturas elegidas")
    rows = []
    for _, row in df_scope.iterrows():
        r = {"Grado": row.get("grado"), "Universidad": row.get("universidad") or "—"}
        for s in chosen:
            v = row.get(s)
            try:
                v = float(v)
            except Exception:
                v = 0.0
            r[s] = round(v, 2) if v in (0.1, 0.2) else 0.0
        rows.append(r)

    out = pd.DataFrame(rows).sort_values(["Grado", "Universidad"]).reset_index(drop=True)
    render_sticky_table(out, sticky_cols=2, height_px=560)

    csv = out.to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇️ Descargar resultados (CSV)",
        data=csv,
        file_name="ponderaciones_4_asignaturas.csv",
        mime="text/csv",
    )

    st.markdown("### 🗂️ Datos completos (Excel) de los grados filtrados")
    with st.expander("Ver todas las columnas del Excel para los grados/universidades seleccionados"):
        all_cols2 = ["grado", "universidad", "titulo_excel"] + SUBJECTS
        st.dataframe(df_scope[all_cols2], use_container_width=True, hide_index=True)
