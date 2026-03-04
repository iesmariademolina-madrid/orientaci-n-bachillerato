# Streamlit app para consultar ponderaciones desde "PONDERACIONES ASIGNATURAS MOLINA.xlsx"
# Versión sin ramas, filtros arriba.
# Flujo: el alumno elige 4 asignaturas (solo entre las que ponderan 0,10 o 0,20) y "descubre" la ponderación.
# v7: La lista de asignaturas ofrecidas se limita estrictamente a las que ponderan 0,10 o 0,20
#     (en lo seleccionado). Se evita mostrar asignaturas con otros valores o vacías.
#
# Ejecuta (Windows recomendado): python -m streamlit run app_orientacion_bachillerato_v7.py

import re
from pathlib import Path
from collections import Counter

import pandas as pd
import openpyxl
import streamlit as st

EXCEL_PATH = "PONDERACIONES ASIGNATURAS MOLINA.xlsx"

st.set_page_config(
    page_title="Orientación Bachillerato · Ponderaciones",
    layout="wide",
    page_icon="🎓",
)

# ---------------------------
# Estilos (color + tabla clara)
# ---------------------------
st.markdown(
    """
<style>
.main {
  background: radial-gradient(circle at 10% 0%, #e0f2fe 0%, transparent 35%),
              radial-gradient(circle at 90% 10%, #ede9fe 0%, transparent 40%),
              radial-gradient(circle at 30% 90%, #dcfce7 0%, transparent 45%),
              linear-gradient(135deg, #f8fafc 0%, #eef2ff 100%);
}
.block-container { padding-top: 1.6rem; }

h1 { color: #0f172a; font-weight: 900; letter-spacing: -0.02em; }
h2, h3 { color: #0b3b8f; }

.stTextInput>div>div>input,
.stMultiSelect>div>div,
.stSelectbox>div>div { border-radius: 12px !important; }

.stButton>button, .stDownloadButton>button {
  background: linear-gradient(90deg, #2563eb 0%, #7c3aed 55%, #db2777 110%);
  color: white !important;
  border: none;
  border-radius: 12px;
  font-weight: 800;
  padding: 0.55rem 0.9rem;
  box-shadow: 0 10px 25px rgba(37, 99, 235, 0.18);
}
.stButton>button:hover, .stDownloadButton>button:hover { filter: brightness(0.98); transform: translateY(-1px); }

.hr {
  height: 1px;
  background: linear-gradient(90deg, transparent, rgba(148, 163, 184, 0.7), transparent);
  margin: 14px 0;
}
.help {
  color: rgba(15, 23, 42, 0.72);
  font-weight: 600;
}
</style>
""",
    unsafe_allow_html=True,
)

# -----------------
# Carga del Excel
# -----------------
@st.cache_data(show_spinner=False)
def load_data(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Asignaturas (fila 1, columnas C..)
    subjects = []
    for c in range(3, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            break
        subjects.append(str(v).strip())

    # Detectar universidades frecuentes entre paréntesis
    deg_cells = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value]
    paren_counter = Counter()
    for s in deg_cells:
        for g in re.findall(r"\(([^()]*)\)", str(s)):
            paren_counter[g.strip()] += 1

    uni_candidates = set()
    for g, cnt in paren_counter.items():
        if (
            cnt >= 5
            and len(g) <= 14
            and re.fullmatch(r"[A-Z0-9/ .-]+", g)
            and (not g.isdigit())
            and (not re.search(r"OPCI|PRESEN|SEMIP|CONJ|INTER|CENTRO|\*", g, re.I))
        ):
            uni_candidates.add(g)
    uni_candidates |= {"UAM", "UCM", "URJC", "UPM", "UC3M", "UAH"}  # fallback

    rows = []
    for r in range(2, ws.max_row + 1):
        full = ws.cell(r, 2).value
        if full is None:
            continue

        s = str(full).strip()
        groups = [g.strip() for g in re.findall(r"\(([^()]*)\)", s)]

        uni = None
        for g in groups:
            if g in uni_candidates:
                uni = g
                break
        if uni is None:
            for g in groups:
                for cand in uni_candidates:
                    if cand in g:
                        uni = cand
                        break
                if uni:
                    break

        grado = re.sub(r"\s*\([^()]*\)\s*", " ", s).strip()
        grado = re.sub(r"\s+", " ", grado)

        coeffs = {}
        for i, subj in enumerate(subjects, start=3):
            v = ws.cell(r, i).value
            if v is None or v == "":
                coeffs[subj] = None
            else:
                try:
                    coeffs[subj] = float(str(v).replace(",", "."))
                except Exception:
                    coeffs[subj] = None

        row = {"grado": grado, "universidad": uni, "titulo_excel": s}
        row.update(coeffs)
        rows.append(row)

    df = pd.DataFrame(rows)
    return df, subjects


# -----------------
# Lógica de filtros
# -----------------
def compute_candidate_subjects_01_02(df_scope: pd.DataFrame, subjects: list[str]) -> list[str]:
    """
    Devuelve SOLO asignaturas que, dentro del ámbito actual (df_scope),
    tienen al menos un 0,10 o un 0,20.
    """
    cand = []
    for s in subjects:
        col = pd.to_numeric(df_scope[s], errors="coerce")
        # tolerancia por flotantes: redondeamos a 2 decimales
        colr = col.round(2)
        if ((colr == 0.10) | (colr == 0.20)).any():
            cand.append(s)
    return cand


def order_subjects_best4(df_scope: pd.DataFrame, subjects: list[str]) -> list[str]:
    """
    Orden robusto para proponer "las 4 mejores" en lo seleccionado.
    Prioriza:
      1) cuántas veces pondera 0,20
      2) cuántas veces pondera 0,10
      3) cuántas veces pondera 0,10 o 0,20 (total)
      4) nombre (estable)
    """
    stats = []
    for s in subjects:
        col = pd.to_numeric(df_scope[s], errors="coerce").round(2)
        if col.dropna().empty:
            continue
        cnt_02 = int((col == 0.20).sum())
        cnt_01 = int((col == 0.10).sum())
        cnt_any = int(((col == 0.20) | (col == 0.10)).sum())
        stats.append((s, cnt_02, cnt_01, cnt_any))

    stats.sort(key=lambda t: (-t[1], -t[2], -t[3], t[0]))
    return [t[0] for t in stats]


def style_weights(df: pd.DataFrame):
    # Colorea celdas por ponderación
    def cell_style(v):
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return "background-color: rgba(148,163,184,0.10); color: rgba(15,23,42,0.55);"
            v = float(v)
        except Exception:
            return ""
        if v == 0:
            return "background-color: rgba(148,163,184,0.10); color: rgba(15,23,42,0.55);"
        if abs(v - 0.2) < 1e-6:
            return "background-color: rgba(16,185,129,0.22); color: #065f46; font-weight: 800;"
        if abs(v - 0.1) < 1e-6:
            return "background-color: rgba(245,158,11,0.22); color: #7c2d12; font-weight: 800;"
        # (por diseño, aquí no deberían aparecer otros valores >0)
        if v > 0:
            return "background-color: rgba(59,130,246,0.18); color: #1e3a8a; font-weight: 800;"
        return ""

    cols = [c for c in df.columns if c not in ("Grado", "Universidad")]
    return df.style.applymap(cell_style, subset=cols).format({c: "{:.2f}" for c in cols})




# -----------------
# Tabla con columnas fijas (Grado/Universidad)
# -----------------
def render_sticky_table(df: pd.DataFrame, height_px: int = 560):
    """
    Renderiza una tabla HTML con "Grado" y "Universidad" fijas al hacer scroll horizontal.
    Streamlit no soporta fijar columnas en st.dataframe, por eso se usa HTML/CSS.
    """
    if df.empty:
        st.info("No hay datos para mostrar.")
        return

    # Convertimos a HTML
    html = df.to_html(index=False, escape=True)

    st.markdown(
        f"""
<style>
.table-wrap {{
  max-height: {height_px}px;
  overflow: auto;
  border-radius: 14px;
  border: 1px solid rgba(148,163,184,0.35);
  background: rgba(255,255,255,0.90);
}}
.table-wrap table {{
  border-collapse: separate;
  border-spacing: 0;
  width: 100%;
  font-size: 0.95rem;
}}
.table-wrap th, .table-wrap td {{
  padding: 8px 10px;
  border-bottom: 1px solid rgba(148,163,184,0.25);
  white-space: nowrap;
}}
.table-wrap th {{
  position: sticky;
  top: 0;
  z-index: 5;
  background: rgba(241,245,249,0.98);
  color: #0f172a;
  font-weight: 900;
}}

/* columnas fijas */
.table-wrap th:nth-child(1), .table-wrap td:nth-child(1) {{
  position: sticky;
  left: 0;
  z-index: 6;
  background: rgba(248,250,252,0.98);
  min-width: 260px;
}}
.table-wrap th:nth-child(2), .table-wrap td:nth-child(2) {{
  position: sticky;
  left: 260px;
  z-index: 6;
  background: rgba(248,250,252,0.98);
  min-width: 140px;
}}
.table-wrap td:nth-child(2), .table-wrap th:nth-child(2) {{
  box-shadow: 10px 0 12px -10px rgba(2,6,23,0.40);
}}

/* colores de ponderación */
.w20 {{ background: rgba(16,185,129,0.22); color: #065f46; font-weight: 900; padding: 2px 6px; border-radius: 8px; }}
.w10 {{ background: rgba(245,158,11,0.22); color: #7c2d12; font-weight: 900; padding: 2px 6px; border-radius: 8px; }}
.w00 {{ background: rgba(148,163,184,0.12); color: rgba(15,23,42,0.65); font-weight: 800; padding: 2px 6px; border-radius: 8px; }}
</style>
""",
        unsafe_allow_html=True,
    )

    # Resaltar 0,20 y 0,10 (tanto con punto como con coma)
    html2 = html
    html2 = re.sub(r">(0\.2(?:0)?|0,2(?:0)?)<", r'><span class="w20">\1</span><', html2)
html2 = re.sub(r">(0\.1(?:0)?|0,1(?:0)?)<", r'><span class="w10">\1</span><', html2)
html2 = re.sub(r">(0\.00|0,00|0)<", r'><span class="w00">0</span><', html2)

    st.markdown(f'<div class="table-wrap">{html2}</div>', unsafe_allow_html=True)



# -----------------
# Notas de corte (PDF DUM)
# -----------------
PDF_CUTOFF_PATH = "DUM NOTAS DE CORTE.pdf"

def _norm(s: str) -> str:
    s = (s or "").lower()
    s = re.sub(r"[áàä]", "a", s)
    s = re.sub(r"[éèë]", "e", s)
    s = re.sub(r"[íìï]", "i", s)
    s = re.sub(r"[óòö]", "o", s)
    s = re.sub(r"[úùü]", "u", s)
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    return s

@st.cache_data(show_spinner=False)
def load_cutoffs(pdf_path: str) -> pd.DataFrame:
    """
    Intenta extraer notas de corte desde el PDF del DUM.
    Devuelve columnas: Universidad, Grado, Nota
    """
    # Primero intentamos pdfplumber (mejor para tablas). Si no está, PyPDF2 como plan B.
    text_pages = []
    try:
        import pdfplumber  # type: ignore
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text() or ""
                if t.strip():
                    text_pages.append(t)
    except Exception:
        try:
            from PyPDF2 import PdfReader  # type: ignore
            reader = PdfReader(pdf_path)
            for p in reader.pages:
                t = p.extract_text() or ""
                if t.strip():
                    text_pages.append(t)
        except Exception as e:
            # Si no se puede leer el PDF, devolvemos vacío (la UI lo explicará)
            return pd.DataFrame(columns=["Universidad", "Grado", "Nota"])

    # Parseo heurístico:
    # - Detecta líneas con "UNIVERSIDAD ..." como cabecera
    # - Detecta filas con "... <nota>" donde la nota tiene coma decimal (p.ej. 11,811)
    rows = []
    current_uni = None

    uni_re = re.compile(r"^\s*UNIVERSIDAD\s+(.+?)\s*$", re.IGNORECASE)
    # nota: 1-2 dígitos, coma, 3 dígitos (en el DUM suele ser así)
    nota_re = re.compile(r"(\d{1,2},\d{3})")

    for page_text in text_pages:
        for line in page_text.splitlines():
            line = line.strip()
            if not line:
                continue

            m_uni = uni_re.match(line)
            if m_uni:
                current_uni = m_uni.group(1).strip()
                continue

            # saltar cabeceras comunes
            if re.search(r"NOTA\s+DE\s+CORTE|TITULACION|PLAZAS|CURSO|DUM", line, re.I):
                continue

            m_nota = nota_re.search(line)
            if m_nota and current_uni:
                nota = m_nota.group(1)
                # grado: texto antes de la nota
                grado = line[: m_nota.start()].strip(" -\t·|")
                # limpiar "Grado en" si aparece
                grado = re.sub(r"^\s*Grado\s+en\s+", "", grado, flags=re.I).strip()
                # evitar líneas demasiado cortas
                if len(grado) >= 3:
                    rows.append({"Universidad": current_uni, "Grado": grado, "Nota": nota})

    dfc = pd.DataFrame(rows).drop_duplicates()
    return dfc

# Mapeo básico (editable) entre códigos de tu Excel y nombres del PDF
UNI_MAP = {
    "UCM": "COMPLUTENSE DE MADRID",
    "UAM": "AUTONOMA DE MADRID",
    "UC3M": "CARLOS III",
    "UPM": "POLITECNICA DE MADRID",
    "URJC": "REY JUAN CARLOS",
    "UAH": "ALCALA",
}


# -----------
# UI principal
# -----------
st.title("🎓 Orientación para elegir materias de Bachillerato")
st.caption("Elige grados/universidades y selecciona 4 asignaturas: verás cómo ponderan en cada opción.")

df, SUBJECTS = load_data(EXCEL_PATH)

with st.container(border=True):
    a1, a2, a3 = st.columns([2.2, 2.2, 1.6])

    with a1:
        q = st.text_input("🔎 Buscar grado (texto)", value="")

    with a2:
        unis = sorted([u for u in df["universidad"].dropna().unique().tolist()])
        b1, b2 = st.columns(2)
        with b1:
            if st.button("🏛️ Seleccionar todas", key="btn_unis_all"):
                st.session_state["sel_unis"] = unis
        with b2:
            if st.button("🧹 Limpiar", key="btn_unis_clear"):
                st.session_state["sel_unis"] = []

        if "sel_unis" not in st.session_state:
            st.session_state["sel_unis"] = []

        selected_unis = st.multiselect("Universidades (opcional)", options=unis, key="sel_unis")

    with a3:
        st.markdown('<div class="help">💡 Consejo: filtra primero por universidades y luego elige grados.</div>', unsafe_allow_html=True)

df_f = df.copy()
if selected_unis:
    df_f = df_f[df_f["universidad"].isin(selected_unis)]
if q.strip():
    df_f = df_f[df_f["grado"].str.lower().str.contains(q.strip().lower(), na=False)]

grados = sorted(df_f["grado"].dropna().unique().tolist())
selected_grados = st.multiselect("📚 Selecciona uno o varios grados", grados)

if not selected_grados:
    st.info("Selecciona al menos un **grado** para ver solo las asignaturas que ponderan en esas opciones.")
    st.stop()

df_scope = df_f[df_f["grado"].isin(selected_grados)]

# ✅ SOLO 0,10 o 0,20
candidate_subjects = compute_candidate_subjects_01_02(df_scope, SUBJECTS)
ordered_subjects = order_subjects_best4(df_scope, candidate_subjects)

# Si cambia el contexto y ya no existen algunas elegidas, las quitamos
if "chosen_subjects" in st.session_state:
    st.session_state["chosen_subjects"] = [s for s in st.session_state["chosen_subjects"] if s in ordered_subjects]

st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

st.subheader("✅ Paso 2: elige 4 asignaturas")
st.caption("Solo aparecen asignaturas que ponderan **0,10 o 0,20** en lo seleccionado. Elige exactamente 4 y descubrirás sus ponderaciones.")

with st.container(border=True):
    b1, b2, b3 = st.columns([1.6, 1.1, 2.3])
    with b1:
        if st.button("✨ Autoseleccionar 4 (mejores)", key="btn_auto4"):
            st.session_state["chosen_subjects"] = ordered_subjects[:4]
    with b2:
        if st.button("🧹 Limpiar selección", key="btn_clear4"):
            st.session_state["chosen_subjects"] = []
    with b3:
        st.caption("Criterio: más veces 0,20; luego más veces 0,10; luego total; y nombre.")

if "chosen_subjects" not in st.session_state:
    st.session_state["chosen_subjects"] = []

chosen = st.multiselect(
    "Asignaturas (elige 4)",
    options=ordered_subjects,
    key="chosen_subjects",
)

if len(chosen) > 4:
    st.warning("Has elegido más de 4. Deja exactamente 4 asignaturas.")
elif len(chosen) < 4:
    st.info(f"Te faltan {4-len(chosen)} asignatura(s) para llegar a 4.")
    st.stop()

st.subheader("📊 Resultado: ponderación de tus 4 asignaturas")

if df_scope.empty:
    st.info("No hay filas con los filtros actuales.")
    st.stop()

rows = []
for _, row in df_scope.iterrows():
    r = {
        "Grado": row.get("grado"),
        "Universidad": row.get("universidad") or "—",
    }
    for s in chosen:
        v = row.get(s)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            r[s] = 0.0
        else:
            try:
                r[s] = float(v)
            except Exception:
                r[s] = 0.0
    rows.append(r)

out = pd.DataFrame(rows).sort_values(["Grado", "Universidad"]).reset_index(drop=True)

render_sticky_table(out, height_px=560)

csv = out.to_csv(index=False).encode("utf-8")
st.download_button(
    "⬇️ Descargar resultados (CSV)",
    data=csv,
    file_name="ponderaciones_4_asignaturas.csv",
    mime="text/csv",
)


# ------------------
# Notas de corte (opcional)
# ------------------
st.markdown("### 🎯 Notas de corte (opcional)")
with st.expander("Ver notas de corte del PDF (DUM NOTAS DE CORTE.pdf) para los grados seleccionados"):
    try:
        df_cut = load_cutoffs(PDF_CUTOFF_PATH)
    except Exception:
        df_cut = pd.DataFrame(columns=["Universidad", "Grado", "Nota"])

    if df_cut.empty:
        st.info("No he podido extraer notas de corte del PDF. Asegúrate de que el archivo 'DUM NOTAS DE CORTE.pdf' está en la misma carpeta y que es texto seleccionable.")
    else:
        # Filtrar por grados seleccionados (matching suave)
        sel_grados_norm = {_norm(g) for g in selected_grados}
        # Map universidad del Excel (código) a fragmento de nombre en PDF
        sel_unis_pdf_frag = []
        for u in selected_unis or []:
            frag = UNI_MAP.get(u, u)
            sel_unis_pdf_frag.append(_norm(frag))

        df_cut["_u"] = df_cut["Universidad"].map(_norm)
        df_cut["_g"] = df_cut["Grado"].map(_norm)

        df_show = df_cut.copy()
        if sel_unis_pdf_frag:
            df_show = df_show[df_show["_u"].apply(lambda x: any(f in x for f in sel_unis_pdf_frag))]

        # grados: si hay selección, filtramos por "contiene" (para tolerar diferencias de nombre)
        if sel_grados_norm:
            df_show = df_show[df_show["_g"].apply(lambda x: any(g in x or x in g for g in sel_grados_norm))]

        df_show = df_show.drop(columns=["_u", "_g"]).reset_index(drop=True)
        st.dataframe(df_show, use_container_width=True, hide_index=True)

        csv_cut = df_show.to_csv(index=False).encode("utf-8")
        st.download_button("⬇️ Descargar notas de corte (CSV)", data=csv_cut, file_name="notas_corte_filtradas.csv", mime="text/csv")

st.markdown("### 🗂️ Datos completos (Excel) de los grados filtrados")
with st.expander("Ver todas las columnas del Excel para los grados/universidades filtrados"):
    all_cols = ["grado", "universidad", "titulo_excel"] + SUBJECTS
    render_sticky_table(df_scope[all_cols].rename(columns={"grado":"Grado","universidad":"Universidad"}), height_px=520)

st.markdown("### 🧾 Notas de corte (si está el PDF en la carpeta)")
with st.expander("Ver nota de corte por grado y universidad (según PDF DUM)"):
    if Path(PDF_CUTOFF_PATH).exists():
        cut = load_cutoffs(PDF_CUTOFF_PATH)
        if cut.empty:
            st.info("No he podido extraer notas de corte del PDF. Si quieres, te lo adapto al formato exacto del documento.")
        else:
            # Filtrar solo a universidades/grados presentes en el ámbito actual
            df_tmp = df_scope[["grado", "universidad"]].copy()
            df_tmp["uni_pdf"] = df_tmp["universidad"].apply(match_uni_label)

            # Emparejar por universidad (PDF) y por grado (coincidencia por texto contenido, case-insensitive)
            # (Más robusto que igualdad exacta)
            cut2 = cut.copy()
            cut2["Grado_lc"] = cut2["Grado"].str.lower()
            rows = []
            for _, r in df_tmp.drop_duplicates().iterrows():
                uni_pdf = str(r["uni_pdf"])
                grado = str(r["grado"])
                grado_lc = grado.lower()
                sub = cut2[cut2["Universidad"].fillna("").str.contains(uni_pdf, case=False, na=False)]
                if sub.empty:
                    continue
                # coincidencias por 'contains' en ambas direcciones
                cand = sub[sub["Grado_lc"].str.contains(grado_lc, na=False) | (pd.Series([grado_lc]*len(sub)).str.contains(sub["Grado_lc"].values))]
                if cand.empty:
                    # intento suave: por primeras 40 letras
                    key = grado_lc[:40]
                    cand = sub[sub["Grado_lc"].str.contains(key, na=False)]
                if cand.empty:
                    continue
                best = cand.sort_values("Nota", ascending=False).iloc[0]
                rows.append({"Grado": grado, "Universidad": r["universidad"] or "—", "Nota de corte": float(best["Nota"])})
            if not rows:
                st.info("No he encontrado coincidencias automáticas entre los nombres de grado del Excel y del PDF. Se puede ajustar con un diccionario de equivalencias.")
            else:
                out_cut = pd.DataFrame(rows).sort_values(["Universidad", "Grado"]).reset_index(drop=True)
                st.dataframe(out_cut, use_container_width=True, hide_index=True)
    else:
        st.warning(f"No encuentro el PDF '{PDF_CUTOFF_PATH}' en la carpeta. Súbelo junto a la app para ver las notas de corte.")

