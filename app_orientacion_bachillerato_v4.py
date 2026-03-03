# Streamlit app para consultar ponderaciones desde "PONDERACIONES ASIGNATURAS MOLINA.xlsx"
# Versión sin ramas, filtros arriba y SIN tarjetas.
# Nuevo flujo: el alumno elige 4 asignaturas (entre las que ponderan) y "descubre" la ponderación por grado/universidad.
#
# Ejecuta (Windows recomendado): python -m streamlit run app_orientacion_bachillerato.py

import re
from collections import Counter

import pandas as pd
import openpyxl
import streamlit as st

EXCEL_PATH = "PONDERACIONES ASIGNATURAS MOLINA.xlsx"

st.set_page_config(
    page_title="Orientación Bachillerato · Ponderaciones",
    layout="wide",
    page_icon="🎓",
)

# ---------------------------
# Estilos (color + tabla clara)
# ---------------------------
st.markdown(
    """
<style>
.main {
  background: radial-gradient(circle at 10% 0%, #e0f2fe 0%, transparent 35%),
              radial-gradient(circle at 90% 10%, #ede9fe 0%, transparent 40%),
              radial-gradient(circle at 30% 90%, #dcfce7 0%, transparent 45%),
              linear-gradient(135deg, #f8fafc 0%, #eef2ff 100%);
}
.block-container { padding-top: 1.6rem; }

h1 { color: #0f172a; font-weight: 900; letter-spacing: -0.02em; }
h2, h3 { color: #0b3b8f; }

.stTextInput>div>div>input,
.stMultiSelect>div>div,
.stSelectbox>div>div { border-radius: 12px !important; }

.stButton>button, .stDownloadButton>button {
  background: linear-gradient(90deg, #2563eb 0%, #7c3aed 55%, #db2777 110%);
  color: white !important;
  border: none;
  border-radius: 12px;
  font-weight: 800;
  padding: 0.55rem 0.9rem;
  box-shadow: 0 10px 25px rgba(37, 99, 235, 0.18);
}
.stButton>button:hover, .stDownloadButton>button:hover { filter: brightness(0.98); transform: translateY(-1px); }

.hr {
  height: 1px;
  background: linear-gradient(90deg, transparent, rgba(148, 163, 184, 0.7), transparent);
  margin: 14px 0;
}
.help {
  color: rgba(15, 23, 42, 0.72);
  font-weight: 600;
}
</style>
""",
    unsafe_allow_html=True,
)

# -----------------
# Carga del Excel
# -----------------
@st.cache_data(show_spinner=False)
def load_data(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Asignaturas (fila 1, columnas C..)
    subjects = []
    for c in range(3, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            break
        subjects.append(str(v).strip())

    # Detectar universidades frecuentes entre paréntesis (URJC, UCM, etc.)
    deg_cells = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value]
    paren_counter = Counter()
    for s in deg_cells:
        for g in re.findall(r"\(([^()]*)\)", str(s)):
            paren_counter[g.strip()] += 1

    uni_candidates = set()
    for g, cnt in paren_counter.items():
        if (
            cnt >= 5
            and len(g) <= 14
            and re.fullmatch(r"[A-Z0-9/ .-]+", g)
            and (not g.isdigit())
            and (not re.search(r"OPCI|PRESEN|SEMIP|CONJ|INTER|CENTRO|\*", g, re.I))
        ):
            uni_candidates.add(g)
    uni_candidates |= {"UAM", "UCM", "URJC", "UPM", "UC3M", "UAH"}  # fallback

    rows = []
    for r in range(2, ws.max_row + 1):
        full = ws.cell(r, 2).value
        if full is None:
            continue

        s = str(full).strip()
        groups = [g.strip() for g in re.findall(r"\(([^()]*)\)", s)]

        uni = None
        for g in groups:
            if g in uni_candidates:
                uni = g
                break
        if uni is None:
            for g in groups:
                for cand in uni_candidates:
                    if cand in g:
                        uni = cand
                        break
                if uni:
                    break

        grado = re.sub(r"\s*\([^()]*\)\s*", " ", s).strip()
        grado = re.sub(r"\s+", " ", grado)

        coeffs = {}
        for i, subj in enumerate(subjects, start=3):
            v = ws.cell(r, i).value
            if v is None or v == "":
                coeffs[subj] = None
            else:
                try:
                    coeffs[subj] = float(str(v).replace(",", "."))
                except Exception:
                    coeffs[subj] = None

        row = {"grado": grado, "universidad": uni, "titulo_excel": s}
        row.update(coeffs)
        rows.append(row)

    df = pd.DataFrame(rows)
    return df, subjects


def compute_candidate_subjects(df_scope: pd.DataFrame, subjects: list[str]) -> list[str]:
    # Solo asignaturas con alguna ponderación > 0 en el ámbito actual
    cand = []
    for s in subjects:
        col = pd.to_numeric(df_scope[s], errors="coerce")
        if (col > 0).any():
            cand.append(s)
    return cand


def order_subjects_consistently(df_scope: pd.DataFrame, subjects: list[str]) -> list[str]:
    # Orden estable: por ponderación máxima desc y nombre
    mx = {}
    for s in subjects:
        m = pd.to_numeric(df_scope[s], errors="coerce").max()
        if pd.isna(m):
            continue
        mx[s] = float(m)
    return sorted(subjects, key=lambda s: (-mx.get(s, 0.0), s))


def style_weights(df: pd.DataFrame):
    # Colorea celdas por ponderación
    def cell_style(v):
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return "background-color: rgba(148,163,184,0.10); color: rgba(15,23,42,0.55);"
            v = float(v)
        except Exception:
            return ""
        if v == 0:
            return "background-color: rgba(148,163,184,0.10); color: rgba(15,23,42,0.55);"
        if abs(v - 0.2) < 1e-9:
            return "background-color: rgba(16,185,129,0.22); color: #065f46; font-weight: 800;"
        if abs(v - 0.1) < 1e-9:
            return "background-color: rgba(245,158,11,0.22); color: #7c2d12; font-weight: 800;"
        if v > 0:
            return "background-color: rgba(59,130,246,0.18); color: #1e3a8a; font-weight: 800;"
        return ""

    cols = [c for c in df.columns if c not in ("Grado", "Universidad")]
    return df.style.applymap(cell_style, subset=cols).format({c: "{:.2f}" for c in cols})


# -----------
# UI principal
# -----------
st.title("🎓 Orientación para elegir materias de Bachillerato")
st.caption("Elige grados/universidades y selecciona 4 asignaturas: verás cómo ponderan en cada opción.")

df, SUBJECTS = load_data(EXCEL_PATH)

with st.container(border=True):
    a1, a2, a3 = st.columns([2.2, 2.2, 1.6])

    with a1:
        q = st.text_input("🔎 Buscar grado (texto)", value="")

    with a2:
        unis = sorted([u for u in df["universidad"].dropna().unique().tolist()])
        b1, b2 = st.columns(2)
        with b1:
            if st.button("🏛️ Seleccionar todas", key="btn_unis_all"):
                st.session_state["sel_unis"] = unis
        with b2:
            if st.button("🧹 Limpiar", key="btn_unis_clear"):
                st.session_state["sel_unis"] = []

        if "sel_unis" not in st.session_state:
            st.session_state["sel_unis"] = []

        selected_unis = st.multiselect(
            "Universidades (opcional)",
            options=unis,
            key="sel_unis",
        )

    with a3:
        st.markdown('<div class="help">💡 Consejo: filtra primero por universidades y luego elige grados.</div>', unsafe_allow_html=True)

df_f = df.copy()
if selected_unis:
    df_f = df_f[df_f["universidad"].isin(selected_unis)]
if q.strip():
    df_f = df_f[df_f["grado"].str.lower().str.contains(q.strip().lower(), na=False)]

grados = sorted(df_f["grado"].dropna().unique().tolist())
selected_grados = st.multiselect("📚 Selecciona uno o varios grados", grados)

df_scope = df_f[df_f["grado"].isin(selected_grados)] if selected_grados else df_f

# Elegir si queremos ver asignaturas que ponderan 0,2 o todas las no nulas
with st.container(border=True):
    m1, m2 = st.columns([1.4, 2.6])
    with m1:
        st.markdown('**🎯 Filtro de asignaturas**')
    with m2:
        if "subj_mode" not in st.session_state:
            st.session_state["subj_mode"] = "all"
        subj_mode_label = st.radio(
            "",
            options=["Todas las que ponderan (>0)", "Solo las que ponderan 0,2"],
            horizontal=True,
            key="subj_mode_label",
        )
        subj_mode = "02" if subj_mode_label == "Solo las que ponderan 0,2" else "all"

# Si no hay grados seleccionados, no mostramos aún el selector de 4 asignaturas
if not selected_grados:
    st.info("Selecciona al menos un **grado** para ver solo las asignaturas que ponderan en esas opciones.")
    st.stop()

# Asignaturas candidatas según el modo elegido
if subj_mode == "02":
    candidate_subjects = []
    for s in SUBJECTS:
        col = pd.to_numeric(df_scope[s], errors="coerce")
        if (col == 0.2).any():
            candidate_subjects.append(s)
else:
    candidate_subjects = compute_candidate_subjects(df_scope, SUBJECTS)

ordered_subjects = order_subjects_consistently(df_scope, candidate_subjects)

st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

st.subheader("✅ Paso 2: elige 4 asignaturas")
st.caption("Verás solo asignaturas que ponderan según el filtro elegido (todas >0 o solo 0,2). Elige exactamente 4 y descubrirás sus ponderaciones.")

# Botones de ayuda para autoselección
with st.container(border=True):
    b1, b2, b3 = st.columns([1.6, 1.1, 2.3])
    with b1:
        if st.button("✨ Autoseleccionar 4 (mejores)", key="btn_auto4"):
            if len(ordered_subjects) >= 4:
                st.session_state["chosen_subjects"] = ordered_subjects[:4]
            else:
                st.session_state["chosen_subjects"] = ordered_subjects
    with b2:
        if st.button("🧹 Limpiar selección", key="btn_clear4"):
            st.session_state["chosen_subjects"] = []
    with b3:
        st.caption("Las 4 se eligen por orden de ponderación máxima en lo seleccionado (y nombre si empatan).")

if "chosen_subjects" not in st.session_state:
    st.session_state["chosen_subjects"] = []

# selector exacto de 4 asignaturas
chosen = st.multiselect(
    "Asignaturas (elige 4)",
    options=ordered_subjects,
    key="chosen_subjects",
)

if len(chosen) > 4:
    st.warning("Has elegido más de 4. Deja exactamente 4 asignaturas.")
elif len(chosen) < 4:
    st.info(f"Te faltan {4-len(chosen)} asignatura(s) para llegar a 4.")
    st.stop()

# ------------------
# Resultados en tabla
# ------------------
st.subheader("📊 Resultado: ponderación de tus 4 asignaturas")
df_sel = df_scope.copy()
if df_sel.empty:
    st.info("No hay filas con los filtros actuales.")
    st.stop()

# construir tabla: una fila por (grado, universidad) del excel
rows = []
for _, row in df_sel.iterrows():
    r = {
        "Grado": row.get("grado"),
        "Universidad": row.get("universidad") or "—",
    }
    for s in chosen:
        v = row.get(s)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            r[s] = 0.0
        else:
            try:
                r[s] = float(v)
            except Exception:
                r[s] = 0.0
    rows.append(r)

out = pd.DataFrame(rows)

# ordenar para que sea legible
out = out.sort_values(["Grado", "Universidad"]).reset_index(drop=True)

# mostrar con estilo
st.dataframe(style_weights(out), use_container_width=True, hide_index=True)

# exportar
csv = out.to_csv(index=False).encode("utf-8")
st.download_button(
    "⬇️ Descargar resultados (CSV)",
    data=csv,
    file_name="ponderaciones_4_asignaturas.csv",
    mime="text/csv",
)

with st.expander("🔎 Ver nombre exacto del título en Excel (opcional)"):
    base_cols = ["grado", "universidad", "titulo_excel"] + chosen
    st.dataframe(df_sel[base_cols], use_container_width=True, hide_index=True)
